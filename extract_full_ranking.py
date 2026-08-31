# -*- coding: utf-8 -*-
"""Generate js/ranking_data.js from the field-test ranking workbook."""

from __future__ import annotations

import json
import re
from datetime import date, datetime
from pathlib import Path
from typing import Any, Iterable

import openpyxl
from openpyxl.worksheet.worksheet import Worksheet


OUTPUT_PATH = Path("js/ranking_data.js")
DATA_SHEET_NAME = "\u30c7\u30fc\u30bf"

DATA_SHEET_VALUE_COLS = {
    "vmax": 42,
    "vdec": 44,
    "sprint_score": 46,
    "pro": 50,
    "dva": 52,
    "eye": 54,
    "peri": 56,
    "flash": 58,
    "arrowz_eye_total": 60,
    "hand_eye": 64,
    "height": 66,
    "weight": 68,
    "bmi": 70,
    "vj": 72,
    "sj": 74,
    "contact_time": 76,
    "jump_height": 78,
    "rj_index": 80,
    "broad_jump": 82,
    "stepping": 84,
}

PERSON_SHEET_VALUE_COLS = {
    "vmax": 22,
    "vdec": 23,
    "sprint_score": 24,
    "pro": 26,
    "dva": 27,
    "eye": 28,
    "peri": 29,
    "flash": 30,
    "arrowz_eye_total": 31,
    "hand_eye": 33,
    "height": 34,
    "weight": 35,
    "bmi": 36,
    "vj": 37,
    "sj": 38,
    "contact_time": 39,
    "jump_height": 40,
    "rj_index": 41,
    "broad_jump": 42,
    "stepping": 43,
}

EXCLUDE_ASCII_PATTERNS = ["Sheet", "CheckList", "TransferLog"]


def find_workbook() -> Path:
    candidates = sorted(Path(".").glob("*.xlsm"), key=lambda p: ("backup" in p.name.lower(), p.name))
    for path in candidates:
        if "backup" not in path.name.lower():
            return path
    if candidates:
        return candidates[0]
    raise FileNotFoundError("No .xlsm workbook found in the project root.")


def parse_date(value: Any) -> datetime | None:
    if value in (None, ""):
        return None
    if isinstance(value, datetime):
        return value
    if isinstance(value, date):
        return datetime(value.year, value.month, value.day)

    text = str(value).strip()
    if not text:
        return None

    for fmt in ("%Y/%m/%d", "%Y-%m-%d", "%Y.%m.%d", "%Y/%m", "%Y-%m"):
        try:
            return datetime.strptime(text, fmt)
        except ValueError:
            pass
    return None


def parse_number(value: Any) -> float | None:
    if value in (None, ""):
        return None
    text = str(value).replace(",", "").strip()
    if text in ("", "-"):
        return None
    try:
        return round(float(text), 4)
    except ValueError:
        return None


def category_from_grade(grade: str) -> str:
    grade = grade.strip()
    match = re.search(r"\d+", grade)

    if "\u5c0f" in grade and match:
        return "U-9" if int(match.group()) <= 3 else "U-12"
    if "\u4e2d" in grade:
        return "U-15"
    if "\u9ad8" in grade:
        return "U-18"
    return "U-12"


def values_from_row(row: tuple[Any, ...], columns: dict[str, int]) -> dict[str, float]:
    scores: dict[str, float] = {}
    for key, index in columns.items():
        if len(row) <= index:
            continue
        parsed = parse_number(row[index])
        if parsed is not None:
            scores[key] = parsed
    return scores


def make_record(
    *,
    name: str,
    class_name: str,
    grade: str,
    gender: str,
    test_date: datetime | None,
    scores: dict[str, float],
) -> dict[str, Any] | None:
    if not name or not scores:
        return None

    return {
        "name": name,
        "class": class_name or grade,
        "category": category_from_grade(grade),
        "grade": grade,
        "gender": gender,
        "test_date": test_date,
        "score": scores.get("vmax", 0),
        "scores": scores,
    }


def read_data_sheet(ws: Worksheet) -> list[dict[str, Any]]:
    records: list[dict[str, Any]] = []

    for row in ws.iter_rows(min_row=2, values_only=True):
        name = str(row[0]).strip() if len(row) > 0 and row[0] not in (None, "") else ""
        if not name or re.match(r"^\d{4}/", name):
            continue

        grade = str(row[11]).strip() if len(row) > 11 and row[11] not in (None, "") else ""
        record = make_record(
            name=name,
            class_name=str(row[9]).strip() if len(row) > 9 and row[9] not in (None, "") else grade,
            grade=grade,
            gender=str(row[10]).strip() if len(row) > 10 and row[10] not in (None, "") else "",
            test_date=parse_date(row[4] if len(row) > 4 else None),
            scores=values_from_row(row, DATA_SHEET_VALUE_COLS),
        )
        if record:
            records.append(record)

    return records


def is_person_sheet(sheet_name: str) -> bool:
    if sheet_name == DATA_SHEET_NAME:
        return False
    return not any(pattern in sheet_name for pattern in EXCLUDE_ASCII_PATTERNS)


def iter_person_sheet_records(sheet_name: str, ws: Worksheet) -> Iterable[dict[str, Any]]:
    if ws.max_column < 44:
        return

    for row in ws.iter_rows(min_row=5, values_only=True):
        if len(row) < 44:
            continue

        test_date = None
        year = parse_number(row[0])
        month = parse_number(row[1])
        if year and month:
            try:
                test_date = datetime(int(year), int(month), 1)
            except ValueError:
                test_date = None
        if test_date is None:
            test_date = parse_date(row[6] if len(row) > 6 else None)
        if test_date is None:
            continue

        grade = str(row[3]).strip() if row[3] not in (None, "") else ""
        record = make_record(
            name=sheet_name.strip(),
            class_name=str(row[7]).strip() if row[7] not in (None, "") else grade,
            grade=grade,
            gender=str(row[4]).strip() if row[4] not in (None, "") else "",
            test_date=test_date,
            scores=values_from_row(row, PERSON_SHEET_VALUE_COLS),
        )
        if record:
            yield record


def read_workbook(workbook_path: Path) -> tuple[str, list[dict[str, Any]]]:
    wb = openpyxl.load_workbook(workbook_path, read_only=True, data_only=True)
    try:
        if DATA_SHEET_NAME in wb.sheetnames:
            records = read_data_sheet(wb[DATA_SHEET_NAME])
            if records:
                return "data sheet", records

        records: list[dict[str, Any]] = []
        for sheet_name in wb.sheetnames:
            if not is_person_sheet(sheet_name):
                continue
            records.extend(iter_person_sheet_records(sheet_name, wb[sheet_name]))
        return "person sheets", records
    finally:
        wb.close()


def keep_latest(records: list[dict[str, Any]]) -> list[dict[str, Any]]:
    latest: dict[str, dict[str, Any]] = {}

    for record in records:
        name = record["name"]
        current = latest.get(name)
        if current is None:
            latest[name] = record
            continue

        new_date = record.get("test_date")
        current_date = current.get("test_date")
        if new_date is not None and (current_date is None or new_date > current_date):
            latest[name] = record

    ranking_data = list(latest.values())
    for item in ranking_data:
        item.pop("test_date", None)
        item.pop("grade", None)
        item.pop("gender", None)

    ranking_data.sort(key=lambda item: item["score"], reverse=True)
    return ranking_data


def write_ranking_data(records: list[dict[str, Any]]) -> None:
    OUTPUT_PATH.parent.mkdir(parents=True, exist_ok=True)
    json_str = json.dumps(records, ensure_ascii=False, separators=(",", ":"))
    OUTPUT_PATH.write_text(f"window.RANKING_DATA = {json_str};\n", encoding="utf-8")


def main() -> None:
    workbook_path = find_workbook()
    source, records = read_workbook(workbook_path)
    ranking_data = keep_latest(records)
    write_ranking_data(ranking_data)

    print(f"Reading file: {workbook_path}")
    print(f"Detected format: {source}")
    print(f"Total records found: {len(records)}")
    print(f"Unique students: {len(ranking_data)}")
    if ranking_data:
        top = ranking_data[0]
        print(f"Top: {top['name']} = {top['score']}")


if __name__ == "__main__":
    main()
